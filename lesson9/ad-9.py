#pip install openpyxl

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

wydatki = [
    ("Czynsz", 1500),
    ("Jedzenie", 800),
    ("Transport", 300)
]

for i, (nazwa, kwota) in enumerate(wydatki, start=1):
    ws[f"A{i}"] = nazwa
    ws[f"B{i}"] = kwota

ws[f"A{len(wydatki)+1}"] = "Suma"
ws[f"B{len(wydatki)+1}"] = f"=SUM(B1:B{len(wydatki)})"

wb.save("finanse.xlsx") 